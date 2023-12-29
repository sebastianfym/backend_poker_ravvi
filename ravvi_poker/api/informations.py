import os

from fastapi import APIRouter
import openpyxl
from ..db import DBI
from ..engine.tables import TablesManager

from .utils import SessionUUID, get_session_and_user

manager = TablesManager()

router = APIRouter(prefix="/info", tags=["info"])


@router.get("/{blinds_type}/{blinds_structure}/blinds_info", status_code=200, summary="Get table (SNG/MTT) result")
async def v1_get_all_info_about_blinds(session_uuid: SessionUUID, blinds_structure: str, blinds_type: str):
    """
    Получаем значения из таблицы о блиндах.

    blinds_type - тип блиндов (sng (можно spinup) или mtt).

    blinds_structure - структура блиндов (стандарт или турбо (для mtt ещще есть гипер турбо)).

    Возвращает список вида [[1, "25/50", 0], [2, "50/75", 1], ...].
    """
    async with DBI() as db:
        _, user = await get_session_and_user(db, session_uuid)

    current_dir = os.path.dirname(os.path.abspath(__file__))

    file_path = os.path.join(current_dir, '..', 'docs', 'blinds', 'blinds_structure.xlsx')

    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook.active

    result_list = []

    if blinds_type.lower() == "sng" or blinds_type.lower() == "spinup":
        if blinds_structure.lower() == "стандарт":
            for row in sheet.iter_rows(min_row=5, max_row=44, min_col=1, max_col=3):
                result_list.append([cell.value for cell in row])

        elif blinds_structure.lower() == "турбо":
            for row in sheet.iter_rows(min_row=5, max_row=40, min_col=5, max_col=7):
                result_list.append([cell.value for cell in row])

    elif blinds_type.lower() == "mtt":
        if blinds_structure.lower() == "стандарт":
            for row in sheet.iter_rows(min_row=5, max_row=84, min_col=9, max_col=11):
                result_list.append([cell.value for cell in row])

        elif blinds_structure.lower() == "турбо":
            for row in sheet.iter_rows(min_row=5, max_row=74, min_col=13, max_col=15):
                result_list.append([cell.value for cell in row])

        elif blinds_structure.lower() == "гипер турбо":
            for row in sheet.iter_rows(min_row=5, max_row=64, min_col=17, max_col=19):
                result_list.append([cell.value for cell in row])

    return result_list


@router.get("/payment structure", status_code=200, summary="Get payment structure")
async def v1_get_payment_structure():
    """
    Получаем значения из таблицы о выигрышах.
    """
    payment_structure_list = [
        {"players": "1-3", "position": {"first": "100%", "second": "0", "third": "0"}},
        {"players": "4-6", "position": {"first": "70%", "second": "30%", "third": "0"}},
        {"players": "7-9", "position": {"first": "50%", "second": "30%", "third": "20%"}}
    ]

    return payment_structure_list


